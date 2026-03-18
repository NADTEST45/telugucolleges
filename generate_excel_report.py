"""
Generate a professional Excel audit report from scraped university data.
Merges results from Selenium and Playwright scrapers.
"""
import json
import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill('solid', fgColor='1F4E79')
HEADER_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=11)
SUBHEADER_FILL = PatternFill('solid', fgColor='D6E4F0')
SUBHEADER_FONT = Font(name='Arial', bold=True, size=10)
DATA_FONT = Font(name='Arial', size=10)
LINK_FONT = Font(name='Arial', size=10, color='0563C1', underline='single')
TITLE_FONT = Font(name='Arial', bold=True, size=14, color='1F4E79')
THIN_BORDER = Border(
    left=Side(style='thin', color='B0B0B0'),
    right=Side(style='thin', color='B0B0B0'),
    top=Side(style='thin', color='B0B0B0'),
    bottom=Side(style='thin', color='B0B0B0'),
)
WRAP_ALIGN = Alignment(wrap_text=True, vertical='top')
CENTER_ALIGN = Alignment(horizontal='center', vertical='center')

STATE_COLORS = {
    'Andhra Pradesh': PatternFill('solid', fgColor='E8F5E9'),
    'Telangana': PatternFill('solid', fgColor='FFF3E0'),
}
TYPE_COLORS = {
    'Deemed University': PatternFill('solid', fgColor='E3F2FD'),
    'State University': PatternFill('solid', fgColor='F3E5F5'),
    'Central University': PatternFill('solid', fgColor='FFF9C4'),
    'Private University': PatternFill('solid', fgColor='FFEBEE'),
}


def merge_results(sel_path='selenium_results.json', pw_path='playwright_results.json'):
    sel_data, pw_data = [], []
    try:
        with open(sel_path, 'r') as f:
            sel_data = json.load(f)
    except FileNotFoundError:
        pass
    try:
        with open(pw_path, 'r') as f:
            pw_data = json.load(f)
    except FileNotFoundError:
        pass

    merged = {}
    for item in sel_data + pw_data:
        key = item['short_name']
        if key not in merged:
            merged[key] = {
                'name': item['name'],
                'short_name': item['short_name'],
                'state': item['state'],
                'city': item['city'],
                'type': item['type'],
                'website': item['website'],
                'all_fees': [],
                'all_courses': [],
                'all_tables': [],
                'errors': [],
            }
        merged[key]['all_fees'].extend(item.get('all_fees', []))
        merged[key]['all_courses'].extend(item.get('all_courses', []))
        merged[key]['all_tables'].extend(item.get('all_tables', []))
        merged[key]['errors'].extend(item.get('errors', []))

    for k in merged:
        merged[k]['all_fees'] = list(dict.fromkeys(merged[k]['all_fees']))
        merged[k]['all_courses'] = sorted(set(merged[k]['all_courses']))

    return list(merged.values())


def apply_header_style(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER


def create_summary_sheet(wb, data):
    ws = wb.active
    ws.title = 'Summary'
    ws.sheet_properties.tabColor = '1F4E79'

    ws.merge_cells('A1:H1')
    ws['A1'] = 'University & Deemed University Audit Report - AP & Telangana'
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:H2')
    ws['A2'] = f'Generated: {datetime.now().strftime("%B %d, %Y %I:%M %p")} | Source: Official University Websites'
    ws['A2'].font = Font(name='Arial', italic=True, size=9, color='666666')
    ws['A2'].alignment = Alignment(horizontal='center')

    headers = ['#', 'University Name', 'Short Name', 'State', 'City',
               'Type', 'Official Website', 'Courses Found', 'Fee Entries Found', 'Data Quality']
    row = 4
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h)
    apply_header_style(ws, row, len(headers))

    for i, uni in enumerate(data, 1):
        r = row + i
        ws.cell(r, 1, i).font = DATA_FONT
        ws.cell(r, 1).alignment = CENTER_ALIGN
        ws.cell(r, 2, uni['name']).font = DATA_FONT
        ws.cell(r, 3, uni['short_name']).font = DATA_FONT
        ws.cell(r, 4, uni['state']).font = DATA_FONT
        if uni['state'] in STATE_COLORS:
            ws.cell(r, 4).fill = STATE_COLORS[uni['state']]
        ws.cell(r, 5, uni['city']).font = DATA_FONT
        ws.cell(r, 6, uni['type']).font = DATA_FONT
        if uni['type'] in TYPE_COLORS:
            ws.cell(r, 6).fill = TYPE_COLORS[uni['type']]
        ws.cell(r, 7, uni['website']).font = LINK_FONT
        ws.cell(r, 8, len(uni['all_courses'])).font = DATA_FONT
        ws.cell(r, 8).alignment = CENTER_ALIGN
        ws.cell(r, 9, len(uni['all_fees'])).font = DATA_FONT
        ws.cell(r, 9).alignment = CENTER_ALIGN

        total = len(uni['all_courses']) + len(uni['all_fees'])
        quality = 'Good' if total > 10 else 'Moderate' if total > 3 else 'Limited'
        ws.cell(r, 10, quality).font = DATA_FONT
        ws.cell(r, 10).alignment = CENTER_ALIGN
        qcolor = {'Good': 'C8E6C9', 'Moderate': 'FFF9C4', 'Limited': 'FFCDD2'}
        ws.cell(r, 10).fill = PatternFill('solid', fgColor=qcolor.get(quality, 'FFFFFF'))

        for c in range(1, len(headers) + 1):
            ws.cell(r, c).border = THIN_BORDER

    widths = [5, 50, 12, 18, 18, 22, 40, 14, 16, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.auto_filter.ref = f'A4:{get_column_letter(len(headers))}{row + len(data)}'
    ws.freeze_panes = 'A5'


def create_fees_sheet(wb, data):
    ws = wb.create_sheet('Fee Details')
    ws.sheet_properties.tabColor = '2E7D32'

    headers = ['University', 'Type', 'State', 'Fee Information (extracted from official website)']
    for col, h in enumerate(headers, 1):
        ws.cell(1, col, h)
    apply_header_style(ws, 1, len(headers))

    r = 2
    for uni in data:
        if not uni['all_fees']:
            ws.cell(r, 1, uni['short_name']).font = DATA_FONT
            ws.cell(r, 2, uni['type']).font = DATA_FONT
            ws.cell(r, 3, uni['state']).font = DATA_FONT
            ws.cell(r, 4, 'No fee data found on official website').font = Font(name='Arial', size=10, italic=True, color='999999')
            for c in range(1, 5):
                ws.cell(r, c).border = THIN_BORDER
            r += 1
        else:
            for fee in uni['all_fees']:
                ws.cell(r, 1, uni['short_name']).font = DATA_FONT
                ws.cell(r, 2, uni['type']).font = DATA_FONT
                ws.cell(r, 3, uni['state']).font = DATA_FONT
                ws.cell(r, 4, fee).font = DATA_FONT
                ws.cell(r, 4).alignment = WRAP_ALIGN
                for c in range(1, 5):
                    ws.cell(r, c).border = THIN_BORDER
                r += 1

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 80
    ws.auto_filter.ref = f'A1:D{r-1}'
    ws.freeze_panes = 'A2'


def create_courses_sheet(wb, data):
    ws = wb.create_sheet('Courses & Programs')
    ws.sheet_properties.tabColor = '1565C0'

    headers = ['University', 'Type', 'State', 'Course / Program (from official website)']
    for col, h in enumerate(headers, 1):
        ws.cell(1, col, h)
    apply_header_style(ws, 1, len(headers))

    r = 2
    for uni in data:
        if not uni['all_courses']:
            ws.cell(r, 1, uni['short_name']).font = DATA_FONT
            ws.cell(r, 2, uni['type']).font = DATA_FONT
            ws.cell(r, 3, uni['state']).font = DATA_FONT
            ws.cell(r, 4, 'No course data found on official website').font = Font(name='Arial', size=10, italic=True, color='999999')
            for c in range(1, 5):
                ws.cell(r, c).border = THIN_BORDER
            r += 1
        else:
            for course in uni['all_courses']:
                ws.cell(r, 1, uni['short_name']).font = DATA_FONT
                ws.cell(r, 2, uni['type']).font = DATA_FONT
                ws.cell(r, 3, uni['state']).font = DATA_FONT
                ws.cell(r, 4, course).font = DATA_FONT
                ws.cell(r, 4).alignment = WRAP_ALIGN
                for c in range(1, 5):
                    ws.cell(r, c).border = THIN_BORDER
                r += 1

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 80
    ws.auto_filter.ref = f'A1:D{r-1}'
    ws.freeze_panes = 'A2'


def create_tables_sheet(wb, data):
    ws = wb.create_sheet('Raw Tables')
    ws.sheet_properties.tabColor = 'F57C00'

    ws.cell(1, 1, 'University').font = HEADER_FONT
    ws.cell(1, 1).fill = HEADER_FILL
    ws.cell(1, 2, 'Table Data (extracted from HTML tables on official sites)').font = HEADER_FONT
    ws.cell(1, 2).fill = HEADER_FILL
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 120

    r = 2
    for uni in data:
        for table in uni.get('all_tables', []):
            for table_row in table:
                ws.cell(r, 1, uni['short_name']).font = DATA_FONT
                ws.cell(r, 2, ' | '.join(str(c) for c in table_row)).font = DATA_FONT
                ws.cell(r, 2).alignment = WRAP_ALIGN
                r += 1
            r += 1
    ws.freeze_panes = 'A2'


def create_errors_sheet(wb, data):
    ws = wb.create_sheet('Scraping Log')
    ws.sheet_properties.tabColor = 'D32F2F'

    headers = ['University', 'Issue']
    for col, h in enumerate(headers, 1):
        ws.cell(1, col, h)
    apply_header_style(ws, 1, len(headers))

    r = 2
    for uni in data:
        for err in uni.get('errors', []):
            ws.cell(r, 1, uni['short_name']).font = DATA_FONT
            ws.cell(r, 2, err).font = DATA_FONT
            ws.cell(r, 2).alignment = WRAP_ALIGN
            r += 1
        if not uni['all_fees'] and not uni['all_courses']:
            ws.cell(r, 1, uni['short_name']).font = DATA_FONT
            ws.cell(r, 2, 'No fee or course data extracted - may need manual review').font = DATA_FONT
            r += 1

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 100
    ws.freeze_panes = 'A2'


def create_stats_sheet(wb, data):
    ws = wb.create_sheet('Statistics')
    ws.sheet_properties.tabColor = '7B1FA2'

    ws.cell(1, 1, 'Audit Statistics').font = TITLE_FONT
    ws.merge_cells('A1:C1')

    stats = [
        ('Total Universities Audited', len(data)),
        ('', ''),
        ('By State:', ''),
        ('  Andhra Pradesh', sum(1 for u in data if u['state'] == 'Andhra Pradesh')),
        ('  Telangana', sum(1 for u in data if u['state'] == 'Telangana')),
        ('', ''),
        ('By Type:', ''),
        ('  Deemed Universities', sum(1 for u in data if u['type'] == 'Deemed University')),
        ('  State Universities', sum(1 for u in data if u['type'] == 'State University')),
        ('  Central Universities', sum(1 for u in data if 'Central' in u['type'])),
        ('  Private Universities', sum(1 for u in data if u['type'] == 'Private University')),
        ('', ''),
        ('Data Coverage:', ''),
        ('  Universities with fee data', sum(1 for u in data if u['all_fees'])),
        ('  Universities with course data', sum(1 for u in data if u['all_courses'])),
        ('  Total fee entries extracted', sum(len(u['all_fees']) for u in data)),
        ('  Total course entries extracted', sum(len(u['all_courses']) for u in data)),
        ('  Universities needing manual review', sum(1 for u in data if not u['all_fees'] and not u['all_courses'])),
    ]

    for i, (label, val) in enumerate(stats, 3):
        ws.cell(i, 1, label).font = SUBHEADER_FONT if label.endswith(':') else DATA_FONT
        if val != '':
            ws.cell(i, 2, val).font = Font(name='Arial', bold=True, size=11)
            ws.cell(i, 2).alignment = CENTER_ALIGN

    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 15


def generate_report(output_path, sel_path='selenium_results.json', pw_path='playwright_results.json'):
    data = merge_results(sel_path, pw_path)
    if not data:
        print("No data to generate report from!")
        return

    data.sort(key=lambda x: (x['state'], x['type'], x['name']))

    wb = Workbook()
    create_summary_sheet(wb, data)
    create_fees_sheet(wb, data)
    create_courses_sheet(wb, data)
    create_tables_sheet(wb, data)
    create_stats_sheet(wb, data)
    create_errors_sheet(wb, data)

    wb.save(output_path)
    print(f"Report saved to {output_path}")
    print(f"  {len(data)} universities processed")
    print(f"  {sum(len(u['all_fees']) for u in data)} fee entries")
    print(f"  {sum(len(u['all_courses']) for u in data)} course entries")
    return output_path


if __name__ == '__main__':
    generate_report('university_audit_report.xlsx')
